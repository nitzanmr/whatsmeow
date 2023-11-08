import subprocess
import re
import openpyxl
import os
# Replace 'your_exe_file.exe' with the actual name of your .exe file
exe_file = "./mdtest.exe"
full_path = os.path.abspath("./")
print(full_path)
# Start the subprocess
process = subprocess.Popen([exe_file], stdout=subprocess.PIPE, text=True, bufsize=1, universal_newlines=True, encoding='utf-8')

# Read and print the output in real-time
while True:
    output_line = process.stdout.readline()
    if output_line is not None:
        print(output_line)
        if 'Received message' in output_line:
        # #Extracting key information using regular expressions
            message_id = re.search(r'Received message (\w+) from', output_line).group(1)
            print(message_id)
            parsed_info = {}
            parsed_info['sender'] = re.search(r'from (\S+)', output_line).group(1)
            parsed_info['pushname'] = re.search(r'pushname: (\w+(?: \w+)?),', output_line).group(1)
            parsed_info['timestamp'] = re.search(r'timestamp: ([\d-]+ [\d:]+ [\+\-\d]+ [A-Z]+)', output_line).group(1)
            parsed_info['message_type'] = re.search(r'type: (\w+)', output_line).group(1)
            if parsed_info['message_type'] == 'text':
                parsed_info['conversation'] = re.search(r'conversation:"([^"]+)"', output_line).group(1)
                parsed_info['media'] = ""
            elif parsed_info['message_type'] == 'media':
                if re.search(r'caption:"([^"]+)"', output_line) is not None:
                    parsed_info['conversation'] = re.search(r'caption:"([^"]+)"', output_line).group(1)
                else:
                    parsed_info['conversation'] = ""
                if(re.search(r'fileName:"([^"]+)', output_line)!= None):
                    full_name = re.search(r'fileName:"([^"]+)', output_line).group(1)
                else:
                    if re.search(r'mimetype:"([^"]+)', output_line).group(1) == "image/jpeg":
                        full_name = message_id + '.jfif'
                    elif re.search(r'mimetype:"([^"]+)', output_line).group(1) == "audio/ogg; codecs=opus":
                        full_name = message_id + ".oga"
                    elif re.search(r'mimetype:"([^"]+)', output_line).group(1) == "video/mp4":
                        full_name = message_id + ".mp4"
                parsed_info['media'] = "file:///" + full_path + "\\" + full_name
            # Define the Excel file and write the data
            excel_file = "whatsapp_output.xlsx"
            # workbook = openpyxl.Workbook()
            # worksheet = workbook.active

            if not os.path.isfile(excel_file):
                # If the file doesn't exist, create a new one with headers
                workbook = openpyxl.Workbook()
                worksheet = workbook.active

                # Set the headers
                headers = parsed_info.keys()
                for col_num, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=1, column=col_num, value=header)
            else:
                # If the file exists, load the existing workbook
                workbook = openpyxl.load_workbook(excel_file)
                worksheet = workbook.active
            # Write the data
            row_data = list(parsed_info.values())
            worksheet.append(row_data)
        #     # Save the Excel file
            workbook.save(excel_file)
        #     # print("Message ID:", message_id)
        #     # print("Sender:", sender)
        #     # print("Pushname:", pushname)
        #     # print("Timestamp:", timestamp)
        #     # print("Message Type:", message_type)
        #     # print("Conversation:", conversation)
    else:
        break

# Wait for the subprocess to complete (optional)
process.wait()


    